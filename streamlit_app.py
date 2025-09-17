import streamlit as st
import pandas as pd
import json
import time
import os
import io
import zipfile
from openai import OpenAI
from datetime import datetime
from Bio import Entrez
import requests
import re
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import xml.etree.ElementTree as ET

# Streamlit UI Configuration
st.set_page_config(
    page_title="GIST",
    page_icon=":mag:",
    layout="wide",
    initial_sidebar_state="expanded",
)

# Helper functions for PDF processing
def convert_pdf_to_txt_file(pdf_file):
    """Convert PDF to a single text file"""
    import PyPDF2
    pdf_reader = PyPDF2.PdfReader(pdf_file)
    text = ""
    for page in pdf_reader.pages:
        text += page.extract_text() + "\n\n"
    return text, len(pdf_reader.pages)


# State Management
if 'analysis_results' not in st.session_state:
    st.session_state['analysis_results'] = []
if 'progress' not in st.session_state:
    st.session_state['progress'] = 0
if 'total_papers' not in st.session_state:
    st.session_state['total_papers'] = 0
if 'search_completed' not in st.session_state:
    st.session_state['search_completed'] = False
if 'no_results_found' not in st.session_state:
    st.session_state['no_results_found'] = False
if 'openai_api_key' not in st.session_state:
    st.session_state['openai_api_key'] = None
if 'api_key_valid' not in st.session_state:
    st.session_state['api_key_valid'] = False
if 'df' not in st.session_state:
    st.session_state['df'] = pd.DataFrame()
if 'pdf_texts' not in st.session_state:
    st.session_state['pdf_texts'] = []
if 'pdf_analysis_completed' not in st.session_state:
    st.session_state['pdf_analysis_completed'] = False
if 'show_clear_confirmation' not in st.session_state:
    st.session_state['show_clear_confirmation'] = False
if 'show_new_search_dialog' not in st.session_state:
    st.session_state['show_new_search_dialog'] = False
if 'search_action' not in st.session_state:
    st.session_state['search_action'] = None
if 'full_text_status' not in st.session_state:
    st.session_state['full_text_status'] = {}
if 'full_query' not in st.session_state:
    st.session_state['full_query'] = ""
if 'unlimited_search' not in st.session_state:
    st.session_state['unlimited_search'] = False

# Function to reset the app state
def reset_app_state():
    st.session_state['analysis_results'] = []
    st.session_state['progress'] = 0
    st.session_state['total_papers'] = 0
    st.session_state['search_completed'] = False
    st.session_state['no_results_found'] = False
    st.session_state['pdf_analysis_completed'] = False
    st.session_state['df'] = pd.DataFrame()
    st.session_state['pdf_texts'] = []
    st.session_state['show_clear_confirmation'] = False
    st.session_state['show_new_search_dialog'] = False
    st.session_state['search_action'] = None
    st.session_state['full_text_status'] = {}
    st.session_state['full_query'] = ""
    st.session_state['unlimited_search'] = False

# Sidebar for Inputs
st.sidebar.header("GIST - Generative Insights Summarization Tool")

# Quick Start Guide dropdown
with st.sidebar.expander("📖 Quick Start Guide"):
    st.markdown("[**🎥 Video Tutorial:**](https://bayergroup.sharepoint.com/:v:/s/LunchandLearnAIandMachineLearning/ER0KU7d2p1NFrrxHqnWUqwEBQsJny-oDDaJulvx3bnj7Vw?e=NxDBdF)")
    
    st.markdown("**📋 How to Use:**")
    st.markdown("""
    1. Enter search terms 
    2. Click ✨ **GIST Analysis**
    3. Download Results: Export findings to Excel format
    """)


# Tab selection
tab_selection = st.sidebar.radio("Select Source", ["PubMed Search", "PDF Upload"])

# OpenAI API Key Input in Sidebar - Now in background
with st.sidebar:
    # Try to authenticate in the background
    if not st.session_state['api_key_valid']:
        try:
            # Get API key from secrets
            openai_api_key = st.secrets["MGA_key"]
            st.session_state['openai_api_key'] = openai_api_key

            # Validate API Key silently
            headers = {
                'accept': 'application/json',
                'x-baychatgpt-accesstoken': openai_api_key
            }
            response = requests.get('https://chat.int.bayer.com/api/v2/users/me', headers=headers)
            response.raise_for_status()
            st.session_state['api_key_valid'] = True
            # No success message displayed
        except KeyError as e:
            # If the key is not found in secrets
            error_message = f"API Key 'MGA_key' not found in secrets. Please enter manually:"
            openai_api_key = st.text_input(error_message, type="password")
        except requests.exceptions.RequestException as e:
            # If API validation fails
            error_message = f"API Key validation failed: {str(e)}. Please enter a valid key:"
            openai_api_key = st.text_input(error_message, type="password")
            
            if openai_api_key:
                st.session_state['openai_api_key'] = openai_api_key
                
                # Validate manually entered API Key
                try:
                    headers = {
                        'accept': 'application/json',
                        'x-baychatgpt-accesstoken': openai_api_key
                    }
                    response = requests.get('https://chat.int.bayer.com/api/v2/users/me', headers=headers)
                    response.raise_for_status()
                    st.session_state['api_key_valid'] = True
                    st.success("API Key Validated! :white_check_mark:")
                except requests.exceptions.RequestException as e:
                    st.error(f"API Key Invalid: {e}")
                    st.session_state['api_key_valid'] = False

    # Disable the analysis buttons if the API key is not provided or invalid
    start_analysis_disabled = not (st.session_state['openai_api_key'] and st.session_state['api_key_valid'])
    
    # Add a clear table button in the sidebar
    st.sidebar.markdown("---")
    if st.sidebar.button("Clear Results Table"):
        st.session_state['show_clear_confirmation'] = True

# Show confirmation popup for clearing the table
if st.session_state['show_clear_confirmation']:
    with st.sidebar:
        st.warning("⚠️ Are you sure you want to clear all results?")
        col1, col2 = st.columns(2)
        if col1.button("Yes, Clear"):
            reset_app_state()
            st.rerun()
        if col2.button("Cancel"):
            st.session_state['show_clear_confirmation'] = False
            st.rerun()

# Configure NCBI Search
if "ncbi_email" in st.secrets:
    Entrez.email = st.secrets["ncbi_email"]
if "PM_Key" in st.secrets:
    Entrez.api_key = st.secrets["PM_Key"]

# Functions
def get_pubmed_count(query):
    """Get the total count of papers available in PubMed for a query."""
    try:
        with Entrez.esearch(db="pubmed", term=query, retmax=0) as handle:
            record = Entrez.read(handle)
        return int(record["Count"])
    except Exception as e:
        return None

def search_and_fetch_pubmed(query, max_results):
    """Search PubMed and fetch details in one function."""
    with Entrez.esearch(db="pubmed", term=query, retmax=max_results) as handle:
        record = Entrez.read(handle)

    id_list = record["IdList"]
    if not id_list:
        return []

    ids = ','.join(id_list)
    with Entrez.efetch(db="pubmed", id=ids, retmode="xml") as handle:
        records = Entrez.read(handle)

    return records

def add_to_excel(output_list, excel_file_path, full_query=""):
    """Add a list of dictionaries to an Excel sheet with formatting."""
    # Create a DataFrame
    df = pd.DataFrame(output_list)
    
    # Define column widths (in Excel units)
    column_widths = {
        "Title": 40,
        "PMID": 10,
        "Full Text Link": 30,
        "Analysis Source": 20,
        "Subject of Study": 15,
        "Disease State": 25,
        "Number of Subjects Studied": 10,
        "Type of Study": 20,
        "Type of Study 2": 15,
        "Study Design": 30,
        "Intervention": 30,
        "Intervention Dose": 30,
        "Intervention Dosage Form": 20,
        "Control": 25,
        "Primary Endpoint": 35,
        "Primary Endpoint Result": 35,
        "Secondary Endpoints": 35,
        "Safety Endpoints": 25,
        "Results Available": 10,
        "Primary Endpoint Met": 10,
        "Statistical Significance": 25,
        "Clinical Significance": 25,
        "Main Author": 20,
        "Other Authors": 40,
        "Journal Name": 30,
        "Date of Publication": 15,
        "Error": 10,
        "Filename": 25,
        "Search Query": 50
    }
    
    # Add the search query to each row if provided
    if full_query:
        df["Search Query"] = full_query
    
    # Save to Excel
    df.to_excel(excel_file_path, index=False, engine='openpyxl')
    
    # Open the Excel file to format it
    wb = openpyxl.load_workbook(excel_file_path)
    ws = wb.active
    
    # Format headers
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Set column widths
    for col_idx, column in enumerate(df.columns, 1):
        col_letter = get_column_letter(col_idx)
        if column in column_widths:
            ws.column_dimensions[col_letter].width = column_widths[column]
        else:
            ws.column_dimensions[col_letter].width = 15  # Default width
    
    # Create a table with the data
    table_ref = f"A1:{get_column_letter(len(df.columns))}{len(df)+1}"
    tab = Table(displayName="ResultsTable", ref=table_ref)
    
    # Add a default style
    style = TableStyleInfo(
        name="TableStyleMedium9", 
        showFirstColumn=False,
        showLastColumn=False, 
        showRowStripes=True, 
        showColumnStripes=False
    )
    tab.tableStyleInfo = style
    
    # Add the table to the worksheet
    ws.add_table(tab)
    
    # Save the formatted workbook
    wb.save(excel_file_path)
    
    return excel_file_path

def analyze_paper(paper, openai_api_key, content_type="abstract"):
    """Call OpenAI to analyze the paper and return the results."""
    client = OpenAI(api_key=openai_api_key, base_url="https://chat.int.bayer.com/api/v2")

    system_prompt = """
    You are a bot speaking with another program that takes JSON formatted text as an input. Only return results in JSON format, with NO PREAMBLE.
    The user will input the results from a PubMed search or a full-text clinical trial PDF. Your job is to extract the exact information to return:
      'Title': The complete article title
      'PMID': The Pubmed ID of the article (if available, otherwise 'NA')
      'Full Text Link' : If available, the DOI URL, otherwise, NA
      'Analysis Source' : Abstract, Full Text, or PDF if explicitly stated
      'Subject of Study': The type of subject in the study. Human, Animal, In-Vitro, Other
      'Disease State': Disease state studied, if any, or if the study is done on a healthy population. leave blank if disease state or healthy patients is not mentioned explicitly. "Healthy patients" if patients are explicitly mentioned to be healthy.
      'Number of Subjects Studied': If human, the total study population. Otherwise, leave blank. This field needs to be an integer or empty.
      'Type of Study': Type of study done. 'RCT' for randomized controlled trial, '1. Meta-analysis','2. Systematic Review','3. Cohort Study', or '4. Other'. If it is '5. Other', append a short description
      'Type of Study 2': Type of clinical study. Evaluate the endpoints - are they clinically relevant endpoints such as symptom relief, or mechanism of action in looking at biomarkers. Output "Clinical" or "MOA" for mechanism of action
      'Study Design': Brief and succinct details about study design, if applicable
      'Intervention': Intervention(s) studied, if any. Intervention is the treatment applied to the group.
      'Intervention Dose': Go in detail here about the intervention's doses and treatment duration if available.
      'Intervention Dosage Form': A brief description of the dosage form - ie. oral, topical, intranasal, if available.
      'Control': Control or comarators, if any
      'Primary Endpoint': What the primary endpoint of the study was, if available. Include how it was measured too if available.
      'Primary Endpoint Result': The measurement for the primary endpoints
      'Secondary Endpoints' If available
      'Safety Endpoints' If available
      'Safety Endpoints Results' Measurements for secondary endpoints, if available. 
      'Results Available': Yes or No
      'Primary Endpoint Met': Summarize from results whether or not the primary endpoint(s) was met: Yes or No or NA if results unavailable
      'Statistical Significance': alpha-level and p-value for primary endpoint(s), if available
      'Clinical Significance': Effect size, and Number needed to treat (NNT)/Number needed to harm (NNH), if available
      'Main Author': Last name, First initials
      'Other Authors': Last name, First initials; Last name First initials; ...
      'Journal Name': Full journal name
      'Date of Publication': YYYY-MM-DD
      'Error': Error description, if any. Otherwise, leave emtpy
    """

    if content_type == "full_text":
        system_prompt += """
        Note: This is a full-text article. Extract as much detail as possible from the full text, focusing on detailed methodology, results, and statistical analyses.
        """
    elif content_type == "pdf":
        system_prompt += """
        Note: This is a full-text PDF of a clinical trial. Extract as much detail as possible from the full text.
        """
    elif content_type == "abstract":
        system_prompt += """
        Note: This is an abstract.
        """

    conversation = client.chat.completions.create(
        model='gpt-4o-mini',  # Use appropriate model
        messages=[
            {
                "role": "system",
                "content": system_prompt
            },
            {
                "role": "user",
                "content": str(paper)
            }
        ]
    )

    cleaned_str = conversation.choices[0].message.content.replace("```json", "").replace("```", "").strip()
    return json.loads(cleaned_str)

# Function to fetch full text from PubMed Central
def fetch_full_text(pmid):
    """Attempt to fetch full text from PubMed Central using the PMID, with PMC/PMID cross-check."""
    try:
        # First, check if the article has a PMC ID
        with Entrez.elink(dbfrom="pubmed", db="pmc", id=pmid) as handle:
            link_results = Entrez.read(handle)
        if not link_results[0]["LinkSetDb"] or not link_results[0]["LinkSetDb"][0]["Link"]:
            return None, "No PMC ID found"
        # Get the PMC ID
        pmc_id = link_results[0]["LinkSetDb"][0]["Link"][0]["Id"]
        # Fetch the full text using the PMC ID
        with Entrez.efetch(db="pmc", id=pmc_id, rettype="xml") as handle:
            full_text_xml = handle.read()
        # Parse XML to extract PMID for cross-check
        try:
            root = ET.fromstring(full_text_xml)
            # Try to find the PMID in the XML (common tag: article-id pub-id-type="pmid")
            pmid_found = None
            for elem in root.iter():
                if elem.tag.endswith("article-id") and elem.attrib.get("pub-id-type") == "pmid":
                    pmid_found = elem.text.strip()
                    break
            if pmid_found is not None and str(pmid_found) != str(pmid):
                # Log a warning and skip
                print(f"Warning: PMC ID {pmc_id} PMID {pmid_found} does not match queried PMID {pmid}. Skipping.")
                return None, f"PMC PMID mismatch: {pmid_found} != {pmid}"
        except Exception as e:
            print(f"Warning: Could not parse PMC XML for PMID cross-check: {e}")
            # If parsing fails, skip for safety
            return None, "PMC XML parse error"
        # Simple extraction of text from XML (could be improved with proper XML parsing)
        text = re.sub(r'<[^>]+>', ' ', full_text_xml.decode('utf-8'))
        text = re.sub(r'\s+', ' ', text).strip()
        return text, "Success"
    except Exception as e:
        return None, f"Error fetching full text: {str(e)}"

# Function to fetch full text from DOI
def fetch_full_text_from_doi(doi):
    """Attempt to fetch full text using DOI through open access APIs."""
    try:
        # Try Unpaywall API
        response = requests.get(f"https://api.unpaywall.org/v2/{doi}?email={Entrez.email}")
        if response.status_code == 200:
            data = response.json()
            if data.get("is_oa") and data.get("best_oa_location") and data.get("best_oa_location").get("url_for_pdf"):
                pdf_url = data["best_oa_location"]["url_for_pdf"]
                pdf_response = requests.get(pdf_url)
                if pdf_response.status_code == 200:
                    # Convert PDF to text
                    pdf_content = io.BytesIO(pdf_response.content)
                    text, _ = convert_pdf_to_txt_file(pdf_content)
                    return text, "Success from DOI"
        
        return None, "Full text not available via DOI"
    
    except Exception as e:
        return None, f"Error fetching from DOI: {str(e)}"

# Function to perform PubMed search and analysis
def perform_pubmed_analysis(query, max_results, action="new"):
    # Store the full query for Excel output
    st.session_state['full_query'] = query
    
    # Reset the no_results_found flag at the beginning of a new search
    st.session_state['no_results_found'] = False
    
    with st.spinner(f"Searching PubMed for '{query}'..."):
        papers = search_and_fetch_pubmed(query, max_results)
        if not papers or 'PubmedArticle' not in papers or not papers['PubmedArticle']:
            st.error("No papers found. Try a different search query.")
            st.session_state['total_papers'] = 0
            st.session_state['no_results_found'] = True
            st.session_state['search_completed'] = True  # Mark as completed but with no results
            return

        st.session_state['total_papers'] = len(papers['PubmedArticle'])
        st.write(f"Found {st.session_state['total_papers']} papers.")

    if st.session_state['total_papers'] > 0:
        progress_bar = st.progress(0)
        
        # Initialize or append to results based on action
        if action == "new":
            st.session_state['analysis_results'] = []
            st.session_state['full_text_status'] = {}
        
        for i, paper in enumerate(papers['PubmedArticle']):
            try:
                with st.spinner(f"Analyzing paper {i+1}/{st.session_state['total_papers']}..."):
                    # First analyze with abstract
                    result = analyze_paper(paper, st.session_state['openai_api_key'], content_type="abstract")
                    
                    # Extract PMID and DOI for full text retrieval
                    pmid = result.get('PMID', 'NA')
                    doi_link = result.get('Full Text Link', 'NA')
                    doi = doi_link.split('/')[-2] + '/' + doi_link.split('/')[-1] if 'doi.org' in doi_link else None
                    
                    # Add source information
                    result['Analysis Source'] = "Abstract"
                    
                    # Store the result
                    st.session_state['analysis_results'].append(result)
                    
                    # Try to get full text
                    full_text = None
                    full_text_source = None
                    
                    # First try PMC
                    if pmid != 'NA':
                        full_text, status = fetch_full_text(pmid)
                        if full_text:
                            full_text_source = "PMC"
                    
                    # If PMC fails, try DOI
                    if not full_text and doi:
                        full_text, status = fetch_full_text_from_doi(doi)
                        if full_text:
                            full_text_source = "DOI"
                    
                    # If we got full text, analyze it and update the result
                    if full_text:
                        with st.spinner(f"Analyzing full text for paper {i+1}..."):
                            full_result = analyze_paper(full_text, st.session_state['openai_api_key'], content_type="full_text")
                            
                            # Preserve the PMID and other identifiers
                            full_result['PMID'] = pmid
                            full_result['Full Text Link'] = doi_link
                            full_result['Analysis Source'] = f"Full Text ({full_text_source})"
                            
                            # Replace the abstract-based result with the full text result
                            st.session_state['analysis_results'][-1] = full_result
                            st.session_state['full_text_status'][pmid] = "Analyzed"
                    else:
                        st.session_state['full_text_status'][pmid] = "Not Available"
                
            except Exception as e:
                st.error(f"Error analyzing paper {i+1}: {e}")

            st.session_state['progress'] = (i + 1) / st.session_state['total_papers']
            progress_bar.progress(st.session_state['progress'])

        st.session_state['search_completed'] = True

# Function to perform PDF analysis
def perform_pdf_analysis(pdf_files, action="new"):
    # Reset the no_results_found flag at the beginning of a new analysis
    st.session_state['no_results_found'] = False
    
    if not pdf_files:
        st.error("No PDF files uploaded. Please upload at least one PDF file.")
        st.session_state['no_results_found'] = True
        st.session_state['search_completed'] = True  # Mark as completed but with no results
        return
    
    st.session_state['total_papers'] = len(pdf_files)
    progress_bar = st.progress(0)
    
    # Initialize or append to results based on action
    if action == "new":
        st.session_state['pdf_texts'] = []
        st.session_state['analysis_results'] = []
    
    for i, pdf_file in enumerate(pdf_files):
        try:
            with st.spinner(f"Processing file {i+1}/{len(pdf_files)}: {pdf_file.name}"):
                # Extract text from PDF
                file_extension = pdf_file.name.split(".")[-1]
                path = pdf_file.read()
                
                if file_extension == "pdf":
                    text_content, _ = convert_pdf_to_txt_file(io.BytesIO(path))
                else:
                    st.error(f"Unsupported file type: {file_extension}. Only PDF files are supported.")
                    continue
                
                # Store extracted text
                st.session_state['pdf_texts'].append({
                    'filename': pdf_file.name,
                    'content': text_content
                })
                
                # Analyze the PDF content
                with st.spinner(f"Analyzing content of {pdf_file.name}..."):
                    result = analyze_paper(text_content, st.session_state['openai_api_key'], content_type="pdf")
                    # Add filename and source to result
                    result['Filename'] = pdf_file.name
                    result['Analysis Source'] = "PDF Full Text"
                    st.session_state['analysis_results'].append(result)
        
        except Exception as e:
            st.error(f"Error processing {pdf_file.name}: {e}")
        
        st.session_state['progress'] = (i + 1) / len(pdf_files)
        progress_bar.progress(st.session_state['progress'])
    
    st.session_state['pdf_analysis_completed'] = True
    st.session_state['search_completed'] = True

# Main UI based on selected tab
if tab_selection == "PubMed Search":
    st.title("PubMed Clinical Trial Analysis")
    
    # PubMed search parameters
    col1, col2 = st.columns([3, 1])
    with col1:
        query = st.text_input("Search Query", "loratadine") + " (clinicaltrial[filter])"
    with col2:
        max_results = st.number_input("Max Results", min_value=1, max_value=400, value=2, step=1)
    
    # Configure file name
    query_word = query.split()[0]  # Extract the first word from the query
    filename = f"Results {query_word} {time.strftime('%y%m%d')}.xlsx"
    
    # Get paper count for both buttons
    paper_count = get_pubmed_count(query)
    if paper_count is not None:
        unlimited_button_text = f"✨ GIST Analysis ({paper_count:,} results)"
        unlimited_help_text = f"Retrieves all {paper_count:,} available results (no limit)"
        
        # For partial analysis, show the minimum of max_results and available papers
        partial_count = min(max_results, paper_count)
        partial_button_text = f"Partial Analysis ({partial_count:,})"
        partial_help_text = f"Retrieves up to {partial_count:,} results (limited by max results setting)"
    else:
        unlimited_button_text = "✨ GIST Analysis"
        unlimited_help_text = "Retrieves all available results (no limit)"
        partial_button_text = "Partial Analysis"
        partial_help_text = f"Retrieves up to {max_results} results (limited by max results setting)"
    
    # Start PubMed analysis buttons
    col1, col2 = st.columns([3, 1])
    
    with col1:
        if st.button(unlimited_button_text, disabled=start_analysis_disabled, help=unlimited_help_text):
            # Check if there are existing results
            if st.session_state['analysis_results'] and len(st.session_state['analysis_results']) > 0:
                st.session_state['show_new_search_dialog'] = True
                st.session_state['unlimited_search'] = True
            else:
                # No existing results, proceed with unlimited search
                perform_pubmed_analysis(query, 10000, action="new")  # Use 10000 as "unlimited"
       
    with col2:
        if st.button(partial_button_text, disabled=start_analysis_disabled, help=partial_help_text):
            # Check if there are existing results
            if st.session_state['analysis_results'] and len(st.session_state['analysis_results']) > 0:
                st.session_state['show_new_search_dialog'] = True
                st.session_state['unlimited_search'] = False
            else:
                # No existing results, proceed with new search
                perform_pubmed_analysis(query, max_results, action="new")

elif tab_selection == "PDF Upload":
    st.title("Clinical Trial PDF Analysis")
    
    # Multiple file uploader
    pdf_files = st.file_uploader("Upload Clinical Trial PDF(s)", type=['pdf'], accept_multiple_files=True)
    
    if pdf_files:
        st.write(f"Uploaded {len(pdf_files)} file(s)")
        
        # Process PDFs button
        if st.button("Process and Analyze PDFs", disabled=start_analysis_disabled):
            # Check if there are existing results
            if st.session_state['analysis_results'] and len(st.session_state['analysis_results']) > 0:
                st.session_state['show_new_search_dialog'] = True
            else:
                # No existing results, proceed with new analysis
                perform_pdf_analysis(pdf_files, action="new")

# Show dialog for new search when results already exist
if st.session_state['show_new_search_dialog']:
    dialog_container = st.container()
    with dialog_container:
        st.warning("### Existing results found")
        st.write("Would you like to start a new search with a fresh table or add to the existing table?")
        
        col1, col2, col3 = st.columns([2, 2, 1])
        
        if col1.button("Start Fresh", key="fresh_button"):
            st.session_state['search_action'] = "new"
            st.session_state['show_new_search_dialog'] = False
            
            # Perform the appropriate action based on the selected tab
            if tab_selection == "PubMed Search":
                search_max_results = 10000 if st.session_state['unlimited_search'] else max_results
                perform_pubmed_analysis(query, search_max_results, action="new")
            else:  # PDF Upload
                perform_pdf_analysis(pdf_files, action="new")
            
            # Force a rerun to refresh the UI and close the dialog
            st.rerun()
        
        if col2.button("Add to Existing", key="add_button"):
            st.session_state['search_action'] = "append"
            st.session_state['show_new_search_dialog'] = False
            
            # Perform the appropriate action based on the selected tab
            if tab_selection == "PubMed Search":
                search_max_results = 10000 if st.session_state['unlimited_search'] else max_results
                perform_pubmed_analysis(query, search_max_results, action="append")
            else:  # PDF Upload
                perform_pdf_analysis(pdf_files, action="append")
            
            # Force a rerun to refresh the UI and close the dialog
            st.rerun()
            
        if col3.button("Cancel", key="cancel_button"):
            st.session_state['show_new_search_dialog'] = False
            st.rerun()
import time
# Display Results and Download Button (common for both tabs)
# Add a new session state variable for tracking the success message timing
if 'success_message_time' not in st.session_state:
    st.session_state['success_message_time'] = None

# Display Results and Download Button (common for both tabs)
if st.session_state['search_completed']:
    if st.session_state['no_results_found']:
        st.warning("No results found. Please try a different search query or upload different files.")
    
    # Always display the table if there are results, regardless of no_results_found flag
    if st.session_state['analysis_results']:
        # Check if we should show the success message
        current_time = time.time()
        
        if st.session_state['success_message_time'] is None:
            # First time showing results, set the timer
            st.session_state['success_message_time'] = current_time
            st.success("Analysis Complete!")
        elif current_time - st.session_state['success_message_time'] < 10:
            # Less than 5 seconds have passed, still show the message
            st.success("Analysis Complete!")
        else:
            # More than 5 seconds have passed, don't show the message
            # No need to rerun
            pass
        
        df = pd.DataFrame(st.session_state['analysis_results'])

        # Convert lists to strings in problematic columns
        for col in df.columns:
            if df[col].dtype == 'object':
                try:
                    df[col] = df[col].astype(str)
                except:
                    print(f"can't convert {col}")
                    pass
        
        # Add a selection column to the dataframe
        df.insert(0, " ", False)
        
        # Store the dataframe in session state
        st.session_state['df'] = df
        
        # Display the interactive table with row selection
        st.write("### Select papers to download:")
        
        # Use the data editor with a checkbox column
        edited_df = st.data_editor(
            df,
            use_container_width=True,
            num_rows="fixed",
            hide_index=True,
            key="results_table"
        )
        
        # Get the selected rows
        selected_df = edited_df[edited_df[" "] == True] if " " in edited_df.columns else pd.DataFrame()
        num_selected = len(selected_df)
        st.write(f"Selected {num_selected} rows")
        
        # File naming based on source
        if tab_selection == "PubMed Search":
            query_word = query.split()[0] if query else "search"
            filename = f"PubMed_{query_word}_{time.strftime('%y%m%d')}.xlsx"
        else:
            filename = f"PDF_Analysis_{time.strftime('%y%m%d')}.xlsx"
        
        # Download buttons
        col1, col2 = st.columns(2)
        
                # Download all results
        df_to_save = df.drop(columns=[" "])
        excel_file_all = add_to_excel(df_to_save.to_dict('records'), filename, st.session_state.get('full_query', ''))
        with open(excel_file_all, "rb") as f:
            col1.download_button(
                label="Download All Results",
                data=f,
                file_name=filename,
                mime="application/vnd.ms-excel",
            )
        
        # Download selected rows
        if num_selected > 0:
            selected_filename = f"Selected_{time.strftime('%y%m%d')}.xlsx"
            selected_excel_file = f"Selected_{filename}"
            selected_df_to_save = selected_df.drop(columns=[" "])
            selected_excel_file = add_to_excel(selected_df_to_save.to_dict('records'), selected_excel_file, st.session_state.get('full_query', ''))
            
            with open(selected_excel_file, "rb") as f:
                col2.download_button(
                    label=f"Download Selected ({num_selected}) Rows",
                    data=f,
                    file_name=selected_filename,
                    mime="application/vnd.ms-excel",
                )
        else:
            col2.write("Select rows to enable partial download")
            
        # If PDF analysis was done, offer text download
        if tab_selection == "PDF Upload" and st.session_state['pdf_analysis_completed']:
            st.write("### Download Extracted Text:")
            
            for i, pdf_text in enumerate(st.session_state['pdf_texts']):
                col1, col2 = st.columns([3, 1])
                col1.write(f"{i+1}. {pdf_text['filename']}")
                col2.download_button(
                    label="Download Text",
                    data=pdf_text['content'],
                    file_name=f"{pdf_text['filename']}.txt",
                    mime="text/plain",
                    key=f"download_text_{i}"
                )

# Footer with styling
st.markdown("""
<style>
.footer {
    position: fixed;
    left: 0;
    bottom: 0;
    width: 100%;
    background-color: rgba(240, 242, 246, 0.9);
    color: #262730;
    text-align: center;
    padding: 10px;
    font-size: 14px;
}
</style>
""", unsafe_allow_html=True)

# Hide Streamlit branding
hide = """
<style>
footer {visibility: hidden;}
.viewerBadge_container__1QSob {visibility: hidden;}
#MainMenu {visibility: hidden;}
</style>
"""
st.markdown(hide, unsafe_allow_html=True)